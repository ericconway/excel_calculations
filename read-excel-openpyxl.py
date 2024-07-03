import openpyxl
from openpyxl.utils import get_column_letter
import random
import json
import os

# Constants
INPUT_CELLS = ['C2', 'C3', 'C4', 'C7', 'C8', 'C9']
CALCULATION_CELLS = ['C12', 'C13']
DEFAULT_FILENAME = 'data.xlsx'
DEFAULT_SHEET_NAME = 'SheetName2'
DEFAULT_ITERATIONS = 10
RANDOM_RANGE = (0.95, 1.05)
DECIMAL_PLACES = 4
OUTPUT_FILENAME = 'calculations-openpyxl.json'

def load_workbook(filename):
    return openpyxl.load_workbook(filename, data_only=False)

def read_cell_value(sheet, cell):
    value = sheet[cell].value
    if isinstance(value, (int, float)):
        return round(value, DECIMAL_PLACES)
    return value

def update_cell_value(sheet, cell, value):
    sheet[cell].value = value

def update_input_randomly(sheet, input_cells):
    for cell in input_cells:
        current_value = read_cell_value(sheet, cell)
        if isinstance(current_value, (int, float)):
            new_value = current_value * random.uniform(*RANDOM_RANGE)
            update_cell_value(sheet, cell, new_value)

def get_all_values(sheet, input_cells, calculation_cells):
    values = {}
    for cell in input_cells + calculation_cells:
        label = sheet[f'A{cell[1:]}'].value
        unit = sheet[f'B{cell[1:]}'].value
        value = read_cell_value(sheet, cell)
        values[label] = {'value': value, 'unit': unit}
    return values

def recalculate_formulas(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                cell.value = cell.value

def main(filename=DEFAULT_FILENAME, sheet_name=DEFAULT_SHEET_NAME, n=DEFAULT_ITERATIONS):
    all_runs = {}
    
        
    for iteration in range(DEFAULT_ITERATIONS):
        wb = load_workbook(filename)
        sheet = wb[sheet_name]
           
        update_input_randomly(sheet, INPUT_CELLS)
            
        # Recalculate formulas
        recalculate_formulas(sheet)
            
        # Save to a temporary file
        temp_filename = f'temp_{filename}'
        wb.save(temp_filename)
            
        # Load the temp file with data_only=True to get calculated values
        wb_data_only = openpyxl.load_workbook(temp_filename, data_only=True)
        sheet_data_only = wb_data_only[sheet_name]
           
        iteration_values = get_all_values(sheet_data_only, INPUT_CELLS, CALCULATION_CELLS)
        all_runs[f"Iteration {iteration}"] = iteration_values
          
        # Save the original workbook with formulas intact
        wb.save(filename)
        
    
    # Save all_runs to calculations.json
    with open(OUTPUT_FILENAME, 'w') as f:
        json.dump(all_runs, f, indent=2)
    
    print(f"Calculations saved to {OUTPUT_FILENAME}")
    
    # Clean up temporary file
    if os.path.exists(temp_filename):
        os.remove(temp_filename)

if __name__ == "__main__":
    main()